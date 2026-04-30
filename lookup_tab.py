"""Lookup tab — full cycle history with search, edit, delete, export."""
import tkinter as tk
from tkinter import ttk, filedialog
from datetime import datetime, timezone, timedelta
from pathlib import Path

import calendar as _calendar

import database
import settings
from utils import to_ist
from ui_helpers import (
    BG, colored_btn, make_upper, scrolled_tree,
    ask_yes_no, show_info, show_error,
)

IST = timedelta(hours=5, minutes=30)

QC_LABELS = {
    'PENDING': 'PENDING',
    'OK':      'OK',
    'NOT_OK':  'NOT OK',
    'LEGACY':  'OK (legacy)',
}

_COLS = ('no', 'rack', 'model', 'qty',
         'smt_op', 'line', 'smt_t',
         'qc_res', 'qc_ins', 'qc_t',
         'th_by', 'th_t')
_HEADS = (
    '#', 'Rack No.', 'Model', 'Cards',
    'SMT Operator', 'Line', 'SMT Handover (IST)',
    'QC Result', 'QC Inspector', 'QC Time (IST)',
    'TH Taken By', 'TH Time (IST)',
)
_WIDTHS = {
    'no': 40, 'qty': 50,
    'line': 80, 'smt_t': 150, 'qc_t': 150, 'th_t': 150,
}
_STRETCH = ('rack', 'model', 'smt_op', 'qc_ins', 'th_by')


def _parse_ist_input(s: str):
    """Parse 'DD/MM/YYYY HH:MM' or 'DD/MM/YYYY' typed by the user → UTC ISO."""
    s = s.strip()
    for fmt in ('%d/%m/%Y %H:%M', '%d/%m/%Y'):
        try:
            naive = datetime.strptime(s, fmt)
            ist_aware = naive.replace(tzinfo=timezone(IST))
            return ist_aware.astimezone(timezone.utc).isoformat()
        except ValueError:
            pass
    return None


def _today_ist_range():
    now_utc = datetime.now(timezone.utc)
    ist_now = now_utc + IST
    start = ist_now.replace(hour=0, minute=0, second=0, microsecond=0) - IST
    end   = ist_now.replace(hour=23, minute=59, second=59, microsecond=0) - IST
    return start.isoformat(), end.isoformat()


def _month_range(year: int, month: int):
    """Return (utc_start, utc_end, 'Month YYYY') for the given IST year/month."""
    last_day = _calendar.monthrange(year, month)[1]
    base = datetime(year, month, 1)
    start_utc = base - IST
    end_utc   = base.replace(day=last_day, hour=23, minute=59, second=59) - IST
    label = base.strftime('%B %Y')
    return start_utc.isoformat(), end_utc.isoformat(), label


def _months_in_range(utc_from: str, utc_to: str):
    """Return list of (year, month) IST months spanned by the UTC range."""
    if not utc_from or not utc_to:
        return []
    try:
        def _to_ist_ym(iso):
            dt = datetime.fromisoformat(iso)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            ist = dt + IST
            return ist.year, ist.month

        y0, m0 = _to_ist_ym(utc_from)
        y1, m1 = _to_ist_ym(utc_to)
        months, y, m = [], y0, m0
        while (y, m) <= (y1, m1):
            months.append((y, m))
            m += 1
            if m > 12:
                m, y = 1, y + 1
        return months
    except Exception:
        return []


class LookupTab:
    def __init__(self, parent):
        self.frame = ttk.Frame(parent, padding=10)
        self.frame.columnconfigure(0, weight=1)
        self._results = []
        self._build()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build(self):
        f = self.frame
        row = 0

        # ── Filter row ────────────────────────────────────────────────────────
        filter_frame = ttk.LabelFrame(f, text='Search Filters', padding=8)
        filter_frame.grid(row=row, column=0, sticky='ew', pady=(0, 6))
        filter_frame.columnconfigure(1, weight=1)
        filter_frame.columnconfigure(3, weight=1)

        ttk.Label(filter_frame, text='RACK NUMBER:',
                  font=('Segoe UI', 10, 'bold')).grid(
            row=0, column=0, sticky='e', padx=(0, 6))
        self._rack_var = tk.StringVar()
        make_upper(self._rack_var)
        ttk.Entry(filter_frame, textvariable=self._rack_var,
                  font=('Segoe UI', 11)).grid(
            row=0, column=1, sticky='ew', ipady=3, padx=(0, 16))

        ttk.Label(filter_frame, text='MODEL:',
                  font=('Segoe UI', 10, 'bold')).grid(
            row=0, column=2, sticky='e', padx=(0, 6))
        self._model_var = tk.StringVar()
        make_upper(self._model_var)
        ttk.Entry(filter_frame, textvariable=self._model_var,
                  font=('Segoe UI', 11)).grid(
            row=0, column=3, sticky='ew', ipady=3)

        # Date range
        ttk.Label(filter_frame, text='FROM (DD/MM/YYYY HH:MM):',
                  font=('Segoe UI', 10, 'bold')).grid(
            row=1, column=0, sticky='e', padx=(0, 6), pady=(6, 0))
        self._from_var = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=self._from_var,
                  font=('Segoe UI', 11)).grid(
            row=1, column=1, sticky='ew', ipady=3, pady=(6, 0), padx=(0, 16))

        ttk.Label(filter_frame, text='TO (DD/MM/YYYY HH:MM):',
                  font=('Segoe UI', 10, 'bold')).grid(
            row=1, column=2, sticky='e', padx=(0, 6), pady=(6, 0))
        self._to_var = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=self._to_var,
                  font=('Segoe UI', 11)).grid(
            row=1, column=3, sticky='ew', ipady=3, pady=(6, 0))

        # Button row
        btn_row = ttk.Frame(filter_frame)
        btn_row.grid(row=2, column=0, columnspan=4, sticky='ew', pady=(8, 0))

        colored_btn(btn_row, 'SEARCH', 'primary',
                    self._on_search, font_size=11, pady=5).pack(side='right', padx=(6, 0))
        colored_btn(btn_row, 'Today', 'secondary',
                    self._set_today, bold=False, pady=5).pack(side='right', padx=(6, 0))
        colored_btn(btn_row, 'All Time', 'secondary',
                    self._set_all_time, bold=False, pady=5).pack(side='right')

        self._set_today()
        row += 1

        # ── Summary bar ───────────────────────────────────────────────────────
        self._summary_var = tk.StringVar(value='No search run yet.')
        tk.Label(f, textvariable=self._summary_var,
                 font=('Segoe UI', 11, 'bold'), bg=BG,
                 fg='#1971c2', anchor='w').grid(
            row=row, column=0, sticky='ew', pady=(0, 1))
        row += 1

        self._model_summary_var = tk.StringVar()
        tk.Label(f, textvariable=self._model_summary_var,
                 font=('Segoe UI', 10), bg=BG,
                 fg='#495057', anchor='w', wraplength=1200, justify='left').grid(
            row=row, column=0, sticky='ew', pady=(0, 4))
        row += 1

        # ── Main table ────────────────────────────────────────────────────────
        container, self._tree = scrolled_tree(
            f, _COLS, _HEADS,
            col_widths=_WIDTHS, stretch_cols=_STRETCH, height=16)
        container.grid(row=row, column=0, sticky='nsew')
        f.rowconfigure(row, weight=1)
        self._tree.bind('<Double-1>', self._on_double_click)
        row += 1

        # ── Bottom action row ─────────────────────────────────────────────────
        bot = ttk.Frame(f)
        bot.grid(row=row, column=0, sticky='ew', pady=(6, 0))

        self._edit_btn = colored_btn(bot, 'Edit Selected', 'primary',
                                     self._on_edit, bold=False, pady=5,
                                     state='disabled')
        self._edit_btn.pack(side='left', padx=(0, 6))

        self._del_btn = colored_btn(bot, 'Delete Selected', 'danger',
                                    self._on_delete, bold=False, pady=5,
                                    state='disabled')
        self._del_btn.pack(side='left')

        self._export_btn = colored_btn(bot, 'Export to Excel', 'success',
                                       self._on_export, font_size=11, pady=5,
                                       state='disabled')
        self._export_btn.pack(side='right')

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _set_today(self):
        now_ist = datetime.now(timezone.utc) + IST
        self._from_var.set(now_ist.strftime('%d/%m/%Y 00:00'))
        self._to_var.set(now_ist.strftime('%d/%m/%Y 23:59'))

    def _set_all_time(self):
        self._from_var.set('01/01/2000 00:00')
        self._to_var.set('31/12/2099 23:59')

    def _selected_cycle(self):
        sel = self._tree.selection()
        if not sel:
            return None
        iid = sel[0]
        idx = self._tree.index(iid)
        if 0 <= idx < len(self._results):
            return self._results[idx]
        return None

    # ── Search ────────────────────────────────────────────────────────────────

    def _on_search(self):
        rack  = self._rack_var.get().strip().upper() or None
        model = self._model_var.get().strip().upper() or None

        utc_from = _parse_ist_input(self._from_var.get())
        utc_to   = _parse_ist_input(self._to_var.get())
        if utc_from is None or utc_to is None:
            show_error('Bad Date',
                       'Use DD/MM/YYYY HH:MM format for date fields (e.g. 09/04/2026 00:00).')
            return

        results = database.get_cycles(rack, utc_from, utc_to, model)
        self._results = results
        self._populate_table(results)

        for btn in (self._edit_btn, self._del_btn, self._export_btn):
            btn.config(state='normal' if results else 'disabled')

        if not results:
            self._summary_var.set('No records found.')
            self._model_summary_var.set('')
        else:
            self._update_summary(results)
            self._update_breakdown(results)

    def _update_summary(self, results):
        pending = sum(1 for r in results if r['qc_result'] == 'PENDING')
        ok_fg   = sum(1 for r in results if r['qc_result'] in ('OK', 'LEGACY') and not r.get('th_id'))
        at_th   = sum(1 for r in results if r.get('th_id'))
        not_ok  = sum(1 for r in results if r['qc_result'] == 'NOT_OK')
        parts   = []
        if pending: parts.append(f"{pending} pending QC")
        if ok_fg:   parts.append(f"{ok_fg} in FG")
        if at_th:   parts.append(f"{at_th} at TH")
        if not_ok:  parts.append(f"{not_ok} returned to SMT")
        self._summary_var.set(
            f"{len(results)} cycle{'s' if len(results) != 1 else ''}  ·  " +
            '  ·  '.join(parts))

    def _update_breakdown(self, results):
        md = {}
        for r in results:
            m = r['model']
            if m not in md:
                md[m] = {'cards': 0, 'pending': 0, 'ok_fg': 0, 'at_th': 0, 'not_ok': 0}
            md[m]['cards'] += settings.resolve_cards(r['quantity'], r.get('cards'), m)
            qr = r['qc_result']
            if qr == 'PENDING':
                md[m]['pending'] += 1
            elif r.get('th_id'):
                md[m]['at_th']  += 1
            elif qr in ('OK', 'LEGACY'):
                md[m]['ok_fg']  += 1
            elif qr == 'NOT_OK':
                md[m]['not_ok'] += 1

        parts = []
        for model in sorted(md):
            d = md[model]
            detail = []
            r = lambda n: f"{n} rack{'s' if n > 1 else ''}"
            if d['at_th']:   detail.append(f"{r(d['at_th'])} to TH")
            if d['ok_fg']:   detail.append(f"{r(d['ok_fg'])} in FG")
            if d['pending']: detail.append(f"{r(d['pending'])} pending")
            if d['not_ok']:  detail.append(f"{r(d['not_ok'])} NOT OK")
            parts.append(f"{model}: {d['cards']} cards  ({',  '.join(detail)})" if detail
                         else f"{model}: {d['cards']} cards")

        self._model_summary_var.set('   |   '.join(parts))

    def _populate_table(self, results):
        self._tree.delete(*self._tree.get_children())
        for i, c in enumerate(results):
            tag = 'odd' if i % 2 == 0 else 'even'
            cards = settings.resolve_cards(c['quantity'], c.get('cards'), c['model'])
            self._tree.insert('', 'end', tags=(tag,), values=(
                i + 1,
                c['rack_number'],
                c['model'],
                cards,
                c.get('smt_operator') or '—',
                c.get('line') or '—',
                to_ist(c['smt_time']) if c.get('smt_time') else '—',
                QC_LABELS.get(c['qc_result'], c['qc_result']),
                c.get('qc_inspector') or '—',
                to_ist(c['qc_time']) if c.get('qc_time') else '—',
                c.get('th_taken_by') or '—',
                to_ist(c['th_time']) if c.get('th_time') else '—',
            ))

    # ── Double-click: show PCB IDs or NOT OK reason ───────────────────────────

    def _on_double_click(self, event):
        col_id = self._tree.identify_column(event.x)
        col_idx = int(col_id.replace('#', '')) - 1
        sel = self._tree.selection()
        if not sel:
            return
        iid = sel[0]
        idx = self._tree.index(iid)
        if idx >= len(self._results):
            return
        cycle = self._results[idx]

        if col_idx == 7:  # QC Result — show NOT OK reason
            reason = cycle.get('not_ok_reason', '')
            if not reason or cycle['qc_result'] != 'NOT_OK':
                return
            dlg = tk.Toplevel(self.frame.winfo_toplevel())
            dlg.title('NOT OK — Reason')
            dlg.resizable(False, False)
            dlg.grab_set()
            frm = ttk.Frame(dlg, padding=16)
            frm.pack(fill='both', expand=True)
            ttk.Label(frm, text='Reason for NOT OK:',
                      font=('Segoe UI', 11, 'bold')).pack(anchor='w')
            tk.Label(frm, text=reason, font=('Segoe UI', 12),
                     bg='#fff5f5', fg='#c92a2a', relief='solid', bd=1,
                     wraplength=380, justify='left', padx=10, pady=8).pack(
                fill='x', pady=8)
            colored_btn(frm, 'Close', 'secondary',
                        dlg.destroy, pady=5).pack(fill='x')
            dlg.wait_window()

    # ── Edit / Delete ─────────────────────────────────────────────────────────

    def _on_edit(self):
        cycle = self._selected_cycle()
        if not cycle:
            return
        from ui_helpers import ask_password
        if not ask_password(self.frame.winfo_toplevel()):
            return
        from edit_dialog import show_edit_dialog
        if show_edit_dialog(dict(cycle), self.frame.winfo_toplevel()):
            self._on_search()

    def _on_delete(self):
        cycle = self._selected_cycle()
        if not cycle:
            return
        from ui_helpers import ask_password
        if not ask_password(self.frame.winfo_toplevel()):
            return
        rack = cycle['rack_number']
        msg = (f"Delete the full cycle for rack {rack}?\n"
               f"This removes the SMT handover and all linked QC/TH records."
               if cycle['cycle_type'] == 'smt' else
               f"Delete this legacy QC/TH record for rack {rack}?")
        if not ask_yes_no('Delete Cycle', msg):
            return
        if cycle['cycle_type'] == 'smt' and cycle.get('smt_id'):
            database.delete_smt_handover(cycle['smt_id'])
        elif cycle.get('ok_scan_id'):
            database.delete_ok_scan(cycle['ok_scan_id'])
        self._on_search()

    # ── Export ────────────────────────────────────────────────────────────────

    def _on_export(self):
        if not self._results:
            return
        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel Files', '*.xlsx')],
            initialfile='rack_cycles_export.xlsx',
            initialdir=str(Path.home()),
        )
        if not path:
            return
        try:
            self._write_excel(path)
            show_info('Export Complete', f"Saved to:\n{path}")
        except Exception as exc:
            show_error('Export Failed', str(exc))

    def _window_status(self, cycle, utc_from, utc_to):
        """Return the cycle's effective status WITHIN the search window.

        TH that happened outside the window is ignored for summary purposes —
        the rack is treated as In FG for that window.
        """
        qr = cycle['qc_result']
        if qr == 'NOT_OK':
            return 'NOT_OK'
        th_time = cycle.get('th_time') or ''
        qc_time = cycle.get('qc_time') or ''
        th_in_window = bool(th_time and utc_from <= th_time <= utc_to)
        if th_in_window:
            return 'TH'
        if qr in ('OK', 'LEGACY'):
            return 'OK_FG'   # QC done in/before window, TH outside window
        return 'PENDING'

    def _write_excel(self, path):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        utc_from = _parse_ist_input(self._from_var.get()) or ''
        utc_to   = _parse_ist_input(self._to_var.get())   or ''

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Rack Cycles'

        HDR_FILL = PatternFill('solid', fgColor='2F5496')
        HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
        NORM_FONT = Font(color='000000', size=11)
        CENTER = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='AAAAAA')
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
        FILLS = {
            'TH':     PatternFill('solid', fgColor='DDEBF7'),
            'OK':     PatternFill('solid', fgColor='E2EFDA'),
            'NOT_OK': PatternFill('solid', fgColor='FADBD8'),
            'PENDING':PatternFill('solid', fgColor='FFF3CD'),
        }
        WHITE = PatternFill('solid', fgColor='FFFFFF')

        col_headers = [
            'Sr. No.', 'Rack Number', 'Model', 'Cards',
            'SMT Operator', 'Line', 'SMT Handover Time (IST)',
            'QC Result', 'NOT OK Reason', 'QC Inspector', 'QC Time (IST)',
            'TH Taken By', 'TH Time (IST)',
        ]
        ws.append(col_headers)
        for col in range(1, len(col_headers) + 1):
            c = ws.cell(1, col)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[1].height = 22

        WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for i, cycle in enumerate(self._results, 1):
            cards = settings.resolve_cards(cycle['quantity'], cycle.get('cards'), cycle['model'])
            ws.append([
                i,
                cycle['rack_number'], cycle['model'], cards,
                cycle.get('smt_operator') or '',
                cycle.get('line') or '',
                to_ist(cycle['smt_time']) if cycle.get('smt_time') else '',
                QC_LABELS.get(cycle['qc_result'], cycle['qc_result']),
                cycle.get('not_ok_reason') or '',
                cycle.get('qc_inspector') or '',
                to_ist(cycle['qc_time']) if cycle.get('qc_time') else '',
                cycle.get('th_taken_by') or '',
                to_ist(cycle['th_time']) if cycle.get('th_time') else '',
            ])

            qr   = cycle['qc_result']
            fill = FILLS['TH'] if cycle.get('th_id') else FILLS.get(qr, WHITE)
            for col in range(1, len(col_headers) + 1):
                c = ws.cell(i + 1, col)
                c.fill = fill; c.font = NORM_FONT
                c.border = BORDER; c.alignment = CENTER
            ws.row_dimensions[i + 1].height = 18

        for col, w in enumerate([8,16,18,8,18,12,24,14,28,18,24,18,24], 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        # Summary sheet
        ws2 = wb.create_sheet('Summary')
        SEC  = Font(bold=True, size=12, color='2F5496')
        TOT  = Font(bold=True, size=11)
        SHDR = PatternFill('solid', fgColor='2F5496')
        SHDR_F = Font(bold=True, color='FFFFFF', size=11)
        ALT  = PatternFill('solid', fgColor='DCE6F1')

        def _sec(t):
            ws2.append([t]); ws2.cell(ws2.max_row, 1).font = SEC
        def _thdr(*cols):
            ws2.append(list(cols)); r = ws2.max_row
            for c in range(1, len(cols)+1):
                cell = ws2.cell(r, c)
                cell.fill = SHDR; cell.font = SHDR_F
                cell.alignment = CENTER; cell.border = BORDER
        def _drow(vals, alt):
            ws2.append(vals); f2 = ALT if alt else WHITE
            for c in range(1, len(vals)+1):
                cell = ws2.cell(ws2.max_row, c)
                cell.fill = f2; cell.font = NORM_FONT
                cell.alignment = CENTER; cell.border = BORDER

        def ws(r): return self._window_status(r, utc_from, utc_to)
        pending = [r for r in self._results if ws(r) == 'PENDING']
        ok_fg   = [r for r in self._results if ws(r) == 'OK_FG']
        at_th   = [r for r in self._results if ws(r) == 'TH']
        not_ok  = [r for r in self._results if ws(r) == 'NOT_OK']

        _sec('OVERALL TOTALS')
        for lbl, val in [('Total Cycles', len(self._results)),
                         ('Pending for QC', len(pending)),
                         ('QC OK — In FG', len(ok_fg)),
                         ('Sent to TH', len(at_th)),
                         ('QC NOT OK — Returned to SMT', len(not_ok))]:
            ws2.append([lbl, val])
            for c in [1,2]:
                cell = ws2.cell(ws2.max_row, c)
                cell.fill = WHITE; cell.font = TOT
                cell.alignment = CENTER; cell.border = BORDER
        ws2.append([])

        _sec('BY MODEL  (status as of search window)')
        _thdr('Model', 'Cycles', 'Pending', 'Pending',
              'OK/FG', 'OK/FG', 'At TH', 'At TH',
              'NOT OK', 'NOT OK', 'Total Cards')
        md = {}
        for r in self._results:
            m = r['model']
            md.setdefault(m, [0, 0,0, 0,0, 0,0, 0,0])
            md[m][0] += 1
            q = settings.resolve_cards(r['quantity'], r.get('cards'), m)
            s = ws(r)
            if s == 'PENDING':  md[m][1] += 1; md[m][2] += q
            elif s == 'OK_FG':  md[m][3] += 1; md[m][4] += q
            elif s == 'TH':     md[m][5] += 1; md[m][6] += q
            elif s == 'NOT_OK': md[m][7] += 1; md[m][8] += q
        for j, (m, v) in enumerate(sorted(md.items()), 1):
            total_cards = v[2] + v[4] + v[6] + v[8]
            _drow([m, v[0], v[1], v[2], v[3], v[4], v[5], v[6], v[7], v[8], total_cards],
                  j % 2 == 0)
        ws2.append([])

        _sec('BY RACK  (status as of search window)')
        _thdr('Rack Number','Cycles','Pending','OK/FG','At TH','NOT OK')
        rd = {}
        for r in self._results:
            rk = r['rack_number']
            rd.setdefault(rk, [0,0,0,0,0])
            rd[rk][0] += 1
            s = ws(r)
            if s == 'PENDING':  rd[rk][1] += 1
            elif s == 'OK_FG':  rd[rk][2] += 1
            elif s == 'TH':     rd[rk][3] += 1
            elif s == 'NOT_OK': rd[rk][4] += 1
        for j, (rk, vals) in enumerate(sorted(rd.items()), 1):
            _drow([rk]+vals, j % 2 == 0)

        for col, w in enumerate([18, 8, 9, 9, 9, 9, 9, 9, 9, 9, 10], 1):
            ws2.column_dimensions[get_column_letter(col)].width = w

        # Monthly summary sheets — one per IST month spanned by the search window
        for yr, mo in _months_in_range(utc_from, utc_to):
            mo_from, mo_to, mo_label = _month_range(yr, mo)
            mo_results = database.get_cycles(utc_from=mo_from, utc_to=mo_to)
            ws3 = wb.create_sheet(mo_label)

            def _sec3(t, _ws=ws3):
                _ws.append([t]); _ws.cell(_ws.max_row, 1).font = SEC
            def _thdr3(*cols, _ws=ws3):
                _ws.append(list(cols)); r3 = _ws.max_row
                for c in range(1, len(cols)+1):
                    cell = _ws.cell(r3, c)
                    cell.fill = SHDR; cell.font = SHDR_F
                    cell.alignment = CENTER; cell.border = BORDER
            def _drow3(vals, alt, _ws=ws3):
                _ws.append(vals); f3 = ALT if alt else WHITE
                for c in range(1, len(vals)+1):
                    cell = _ws.cell(_ws.max_row, c)
                    cell.fill = f3; cell.font = NORM_FONT
                    cell.alignment = CENTER; cell.border = BORDER

            def _ws3_status(r, _f=mo_from, _t=mo_to):
                return self._window_status(r, _f, _t)

            mo_pending = [r for r in mo_results if _ws3_status(r) == 'PENDING']
            mo_ok_fg   = [r for r in mo_results if _ws3_status(r) == 'OK_FG']
            mo_at_th   = [r for r in mo_results if _ws3_status(r) == 'TH']
            mo_not_ok  = [r for r in mo_results if _ws3_status(r) == 'NOT_OK']

            _sec3('OVERALL TOTALS')
            for lbl, val in [('Total Cycles', len(mo_results)),
                             ('Pending for QC', len(mo_pending)),
                             ('QC OK — In FG', len(mo_ok_fg)),
                             ('Sent to TH', len(mo_at_th)),
                             ('QC NOT OK — Returned to SMT', len(mo_not_ok))]:
                ws3.append([lbl, val])
                for c in [1, 2]:
                    cell = ws3.cell(ws3.max_row, c)
                    cell.fill = WHITE; cell.font = TOT
                    cell.alignment = CENTER; cell.border = BORDER
            ws3.append([])

            _sec3('BY MODEL')
            _thdr3('Model', 'Cycles', 'Pending', 'Pending',
                   'OK/FG', 'OK/FG', 'At TH', 'At TH',
                   'NOT OK', 'NOT OK', 'Total Cards')
            mmd = {}
            for r in mo_results:
                m = r['model']
                mmd.setdefault(m, [0, 0,0, 0,0, 0,0, 0,0])
                mmd[m][0] += 1
                q = settings.resolve_cards(r['quantity'], r.get('cards'), m)
                s = _ws3_status(r)
                if s == 'PENDING':  mmd[m][1] += 1; mmd[m][2] += q
                elif s == 'OK_FG':  mmd[m][3] += 1; mmd[m][4] += q
                elif s == 'TH':     mmd[m][5] += 1; mmd[m][6] += q
                elif s == 'NOT_OK': mmd[m][7] += 1; mmd[m][8] += q
            for j, (m, v) in enumerate(sorted(mmd.items()), 1):
                total_cards = v[2] + v[4] + v[6] + v[8]
                _drow3([m, v[0], v[1], v[2], v[3], v[4], v[5], v[6], v[7], v[8], total_cards],
                       j % 2 == 0)
            ws3.append([])

            _sec3('BY RACK')
            _thdr3('Rack Number', 'Cycles', 'Pending', 'OK/FG', 'At TH', 'NOT OK')
            rrd = {}
            for r in mo_results:
                rk = r['rack_number']
                rrd.setdefault(rk, [0,0,0,0,0])
                rrd[rk][0] += 1
                s = _ws3_status(r)
                if s == 'PENDING':  rrd[rk][1] += 1
                elif s == 'OK_FG':  rrd[rk][2] += 1
                elif s == 'TH':     rrd[rk][3] += 1
                elif s == 'NOT_OK': rrd[rk][4] += 1
            for j, (rk, vals) in enumerate(sorted(rrd.items()), 1):
                _drow3([rk]+vals, j % 2 == 0)

            for col, w in enumerate([18, 8, 9, 9, 9, 9, 9, 9, 9, 9, 10], 1):
                ws3.column_dimensions[get_column_letter(col)].width = w

        wb.save(path)

    def on_activate(self):
        self._rack_var.set('')
